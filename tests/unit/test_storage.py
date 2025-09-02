"""
Tests for StorageService.
"""
import pytest
import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
import hashlib
from openpyxl import load_workbook

from services.storage import StorageService


class TestStorageService:
    """Test cases for StorageService."""
    
    def test_parquet_round_trip(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test writing and reading parquet files with data integrity."""
        parquet_path = temp_dir / "test.parquet"
        
        # Write parquet
        checksum = StorageService.write_parquet(sample_df, parquet_path)
        
        # Verify file exists and checksum is returned
        assert parquet_path.exists()
        assert isinstance(checksum, str)
        assert len(checksum) == 64  # SHA256 hex length
        
        # Read parquet and verify data integrity
        read_df = StorageService.read_parquet(parquet_path)
        
        pd.testing.assert_frame_equal(sample_df, read_df)
    
    def test_parquet_round_trip_with_selected_columns(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test reading specific columns from parquet."""
        parquet_path = temp_dir / "test.parquet"
        
        StorageService.write_parquet(sample_df, parquet_path)
        
        # Read only specific columns
        columns_to_read = ["Company", "Revenue"]
        read_df = StorageService.read_parquet(parquet_path, columns=columns_to_read)
        
        expected_df = sample_df[columns_to_read]
        pd.testing.assert_frame_equal(expected_df, read_df)
    
    def test_parquet_with_various_dtypes(self, temp_dir: Path):
        """Test parquet round-trip with various data types."""
        df_with_types = pd.DataFrame({
            "string_col": ["A", "B", "C"],
            "int_col": [1, 2, 3],
            "float_col": [1.1, 2.2, 3.3],
            "bool_col": [True, False, True],
            "date_col": pd.date_range("2023-01-01", periods=3)
        })
        
        parquet_path = temp_dir / "types_test.parquet"
        StorageService.write_parquet(df_with_types, parquet_path)
        read_df = StorageService.read_parquet(parquet_path)
        
        pd.testing.assert_frame_equal(df_with_types, read_df)
    
    def test_csv_write_read(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test CSV write and read operations."""
        csv_path = temp_dir / "test.csv"
        
        # Write CSV
        result_path = StorageService.write_csv(sample_df, csv_path)
        assert result_path == str(csv_path)
        assert csv_path.exists()
        
        # Read CSV and verify
        read_df = StorageService.read_csv(csv_path)
        pd.testing.assert_frame_equal(sample_df, read_df)
    
    def test_csv_with_custom_params(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test CSV operations with custom parameters."""
        csv_path = temp_dir / "test_custom.csv"
        
        # Write with custom separator
        StorageService.write_csv(sample_df, csv_path, sep=";")
        
        # Read with matching separator
        read_df = StorageService.read_csv(csv_path, sep=";")
        pd.testing.assert_frame_equal(sample_df, read_df)
    
    def test_excel_write_single_sheet(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test writing single sheet Excel file."""
        excel_path = temp_dir / "test_single.xlsx"
        
        result_path = StorageService.write_excel(sample_df, excel_path)
        assert result_path == str(excel_path)
        assert excel_path.exists()
        
        # Read and verify
        read_df = StorageService.read_excel(excel_path)
        pd.testing.assert_frame_equal(sample_df, read_df)
    
    def test_excel_write_multiple_sheets(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test writing Excel file with multiple sheets."""
        excel_path = temp_dir / "test_multi.xlsx"
        
        # Create sample data for multiple sheets
        sheets_data = {
            "Revenue": sample_df,
            "Summary": pd.DataFrame({
                "Total_Revenue": [sample_df["Revenue"].sum()],
                "Company_Count": [len(sample_df)]
            })
        }
        
        result_path = StorageService.write_excel(sheets_data, excel_path)
        assert result_path == str(excel_path)
        assert excel_path.exists()
        
        # Verify both sheets exist and contain correct data
        workbook = load_workbook(excel_path)
        assert "Revenue" in workbook.sheetnames
        assert "Summary" in workbook.sheetnames
        
        # Read and verify each sheet
        revenue_df = StorageService.read_excel(excel_path, sheet_name="Revenue")
        pd.testing.assert_frame_equal(sample_df, revenue_df)
        
        summary_df = StorageService.read_excel(excel_path, sheet_name="Summary")
        expected_summary = sheets_data["Summary"]
        pd.testing.assert_frame_equal(expected_summary, summary_df)
    
    def test_excel_read_specific_sheet(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test reading specific sheet from Excel."""
        excel_path = temp_dir / "test_sheets.xlsx"
        
        sheets_data = {
            "Sheet1": sample_df,
            "Sheet2": pd.DataFrame({"Col1": [1, 2, 3]})
        }
        
        StorageService.write_excel(sheets_data, excel_path)
        
        # Read specific sheet
        read_df = StorageService.read_excel(excel_path, sheet_name="Sheet2")
        pd.testing.assert_frame_equal(sheets_data["Sheet2"], read_df)
    
    def test_calculate_checksum(self, temp_dir: Path):
        """Test SHA256 checksum calculation."""
        test_file = temp_dir / "checksum_test.txt"
        test_content = b"Hello, World!"
        
        # Write test content
        with open(test_file, "wb") as f:
            f.write(test_content)
        
        # Calculate checksum
        checksum = StorageService.calculate_checksum(test_file)
        
        # Verify against manual calculation
        expected = hashlib.sha256(test_content).hexdigest()
        assert checksum == expected
        assert len(checksum) == 64
    
    def test_calculate_checksum_large_file(self, temp_dir: Path):
        """Test checksum calculation for larger files (tests chunked reading)."""
        test_file = temp_dir / "large_test.txt"
        
        # Create content larger than read buffer (4096 bytes)
        test_content = b"A" * 10000
        
        with open(test_file, "wb") as f:
            f.write(test_content)
        
        checksum = StorageService.calculate_checksum(test_file)
        expected = hashlib.sha256(test_content).hexdigest()
        assert checksum == expected
    
    def test_save_upload_with_checksum(self, temp_dir: Path):
        """Test saving uploaded file content with checksum verification."""
        upload_path = temp_dir / "subdir" / "uploaded_file.bin"
        test_content = b"This is uploaded file content"
        
        # Save upload (should create parent directories)
        checksum = StorageService.save_upload(test_content, upload_path)
        
        # Verify file exists and parent directory was created
        assert upload_path.exists()
        assert upload_path.parent.exists()
        
        # Verify content
        with open(upload_path, "rb") as f:
            saved_content = f.read()
        assert saved_content == test_content
        
        # Verify checksum
        expected_checksum = hashlib.sha256(test_content).hexdigest()
        assert checksum == expected_checksum
    
    def test_empty_dataframe_operations(self, temp_dir: Path):
        """Test operations with empty DataFrames."""
        empty_df = pd.DataFrame()
        
        # Test parquet
        parquet_path = temp_dir / "empty.parquet"
        checksum = StorageService.write_parquet(empty_df, parquet_path)
        assert isinstance(checksum, str)
        
        read_df = StorageService.read_parquet(parquet_path)
        assert len(read_df) == 0
        
        # Test CSV (empty DataFrames create CSV files that can't be read back)
        csv_path = temp_dir / "empty.csv"
        StorageService.write_csv(empty_df, csv_path)
        # For empty DataFrames, CSV reading fails - this is expected pandas behavior
        # We'll just verify the file exists
        assert csv_path.exists()
        
        # Test Excel
        excel_path = temp_dir / "empty.xlsx"
        StorageService.write_excel(empty_df, excel_path)
        read_df = StorageService.read_excel(excel_path)
        assert len(read_df) == 0
    
    def test_special_characters_handling(self, temp_dir: Path):
        """Test handling of special characters in data."""
        special_df = pd.DataFrame({
            "Name": ["John Döe", "María García", "李小明"],
            "Notes": ["Special chars: áéíóú", "Currency: $€¥", "Symbols: @#%&*"],
            "Unicode": ["🚀", "🎉", "✨"]
        })
        
        # Test parquet
        parquet_path = temp_dir / "special.parquet"
        StorageService.write_parquet(special_df, parquet_path)
        read_df = StorageService.read_parquet(parquet_path)
        pd.testing.assert_frame_equal(special_df, read_df)
        
        # Test CSV
        csv_path = temp_dir / "special.csv"
        StorageService.write_csv(special_df, csv_path)
        read_df = StorageService.read_csv(csv_path)
        pd.testing.assert_frame_equal(special_df, read_df)
        
        # Test Excel
        excel_path = temp_dir / "special.xlsx"
        StorageService.write_excel(special_df, excel_path)
        read_df = StorageService.read_excel(excel_path)
        pd.testing.assert_frame_equal(special_df, read_df)
    
    def test_path_as_string_and_pathlib(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test that methods accept both string and Path objects."""
        # Test with pathlib.Path
        parquet_path_obj = temp_dir / "test_path.parquet"
        checksum1 = StorageService.write_parquet(sample_df, parquet_path_obj)
        
        # Test with string path
        parquet_path_str = str(temp_dir / "test_string.parquet")
        checksum2 = StorageService.write_parquet(sample_df, parquet_path_str)
        
        # Both should work
        assert isinstance(checksum1, str)
        assert isinstance(checksum2, str)
        assert Path(parquet_path_obj).exists()
        assert Path(parquet_path_str).exists()
    
    def test_dataframe_with_null_values(self, temp_dir: Path):
        """Test handling DataFrames with null values."""
        df_with_nulls = pd.DataFrame({
            "Company": ["ACME", None, "Beta"],
            "Revenue": [1000, 2000, None],
            "Notes": ["Good", "", np.nan]
        })
        
        # Test parquet (preserves nulls better)
        parquet_path = temp_dir / "nulls.parquet"
        StorageService.write_parquet(df_with_nulls, parquet_path)
        read_df = StorageService.read_parquet(parquet_path)
        
        # Check that nulls are preserved (might need to handle NaN comparison carefully)
        assert pd.isna(read_df.loc[1, "Company"])
        assert pd.isna(read_df.loc[2, "Revenue"])
        assert pd.isna(read_df.loc[2, "Notes"])
    
    def test_write_parquet_with_kwargs(self, temp_dir: Path, sample_df: pd.DataFrame):
        """Test parquet writing with additional kwargs."""
        parquet_path = temp_dir / "test_kwargs.parquet"
        
        # Test with compression parameter
        checksum = StorageService.write_parquet(
            sample_df, 
            parquet_path, 
            compression="snappy"
        )
        
        assert isinstance(checksum, str)
        assert parquet_path.exists()
        
        # Verify data integrity
        read_df = StorageService.read_parquet(parquet_path)
        pd.testing.assert_frame_equal(sample_df, read_df)