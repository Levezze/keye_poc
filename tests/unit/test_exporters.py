"""
Tests for ExportService.
"""
import pytest
import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, Any

from services.exporters import ExportService


class TestExportService:
    """Test cases for ExportService."""
    
    def test_export_concentration_csv(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test CSV export with correct rows per (period, threshold)."""
        csv_path = temp_dir / "concentration_results.csv"
        
        result_path = ExportService.export_concentration_csv(
            sample_concentration_results,
            csv_path
        )
        
        assert result_path == str(csv_path)
        assert csv_path.exists()
        
        # Read and verify CSV structure
        df = pd.read_csv(csv_path)
        
        # Verify columns
        expected_columns = ["period", "threshold", "count", "value", "pct_of_total"]
        assert list(df.columns) == expected_columns
        
        # Verify row count: 2 periods × 3 thresholds = 6 rows
        assert len(df) == 6
        
        # Verify data for first period
        q1_data = df[df["period"] == "2023-Q1"]
        assert len(q1_data) == 3  # 3 thresholds
        
        # Check specific threshold data
        q1_top10 = q1_data[q1_data["threshold"] == 10].iloc[0]
        assert q1_top10["count"] == 1
        assert q1_top10["value"] == 1000000
        assert q1_top10["pct_of_total"] == 40.0
        
        q1_top20 = q1_data[q1_data["threshold"] == 20].iloc[0]
        assert q1_top20["count"] == 2
        assert q1_top20["value"] == 1750000
        assert q1_top20["pct_of_total"] == 70.0
        
        q1_top50 = q1_data[q1_data["threshold"] == 50].iloc[0]
        assert q1_top50["count"] == 4
        assert q1_top50["value"] == 2500000
        assert q1_top50["pct_of_total"] == 100.0
        
        # Verify data for second period
        q2_data = df[df["period"] == "2023-Q2"]
        assert len(q2_data) == 3  # 3 thresholds
        
        q2_top10 = q2_data[q2_data["threshold"] == 10].iloc[0]
        assert q2_top10["count"] == 1
        assert q2_top10["value"] == 800000
        assert q2_top10["pct_of_total"] == 40.0
    
    def test_export_concentration_csv_missing_by_period(self, temp_dir: Path):
        """Test CSV export with missing by_period data."""
        csv_path = temp_dir / "empty_concentration.csv"
        results = {"group_by": "Company", "value_column": "Revenue"}
        
        result_path = ExportService.export_concentration_csv(results, csv_path)
        assert result_path == str(csv_path)
        assert csv_path.exists()
        
        # When there's no data, an empty DataFrame with no columns is created
        # This is the expected behavior for this edge case
        try:
            df = pd.read_csv(csv_path)
            assert len(df) == 0  # Empty DataFrame
        except pd.errors.EmptyDataError:
            # This is expected when the CSV has no columns
            pass
    
    def test_export_concentration_excel_sheets_present(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test Excel export has all expected sheets (Summary, Details, Parameters)."""
        excel_path = temp_dir / "concentration_results.xlsx"
        
        result_path = ExportService.export_concentration_excel(
            sample_concentration_results,
            excel_path
        )
        
        assert result_path == str(excel_path)
        assert excel_path.exists()
        
        # Load workbook and check sheets
        workbook = load_workbook(excel_path)
        sheet_names = workbook.sheetnames
        
        assert "Summary" in sheet_names
        assert "Details" in sheet_names  
        assert "Parameters" in sheet_names
    
    def test_export_concentration_excel_summary_sheet(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test Excel Summary sheet contains all threshold metrics."""
        excel_path = temp_dir / "concentration_summary.xlsx"
        
        ExportService.export_concentration_excel(sample_concentration_results, excel_path)
        
        # Read summary sheet
        summary_df = pd.read_excel(excel_path, sheet_name="Summary")
        
        # Check columns exist for all thresholds
        expected_columns = [
            "period", "total",
            "top_10_count", "top_10_value", "top_10_pct",
            "top_20_count", "top_20_value", "top_20_pct",
            "top_50_count", "top_50_value", "top_50_pct"
        ]
        
        for col in expected_columns:
            assert col in summary_df.columns, f"Missing column: {col}"
        
        # Verify data for first period
        q1_row = summary_df[summary_df["period"] == "2023-Q1"].iloc[0]
        assert q1_row["total"] == 2500000
        assert q1_row["top_10_count"] == 1
        assert q1_row["top_10_value"] == 1000000
        assert q1_row["top_10_pct"] == 40.0
        assert q1_row["top_20_count"] == 2
        assert q1_row["top_20_value"] == 1750000
        assert q1_row["top_20_pct"] == 70.0
        assert q1_row["top_50_count"] == 4
        assert q1_row["top_50_value"] == 2500000
        assert q1_row["top_50_pct"] == 100.0
    
    def test_export_concentration_excel_details_sheet(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test Excel Details sheet contains correct data."""
        excel_path = temp_dir / "concentration_details.xlsx"
        
        ExportService.export_concentration_excel(sample_concentration_results, excel_path)
        
        # Read details sheet
        details_df = pd.read_excel(excel_path, sheet_name="Details")
        
        # Verify details match input
        assert len(details_df) == 2
        assert "Company" in details_df.columns
        assert "Total_Revenue" in details_df.columns
        assert "Rank" in details_df.columns
        
        # Check specific data
        acme_row = details_df[details_df["Company"] == "ACME Corp"].iloc[0]
        assert acme_row["Total_Revenue"] == 1800000
        assert acme_row["Rank"] == 1
        
        beta_row = details_df[details_df["Company"] == "Beta Inc"].iloc[0]
        assert beta_row["Total_Revenue"] == 1200000
        assert beta_row["Rank"] == 2
    
    def test_export_concentration_excel_parameters_sheet(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test Excel Parameters sheet contains configuration data."""
        excel_path = temp_dir / "concentration_params.xlsx"
        
        ExportService.export_concentration_excel(sample_concentration_results, excel_path)
        
        # Read parameters sheet
        params_df = pd.read_excel(excel_path, sheet_name="Parameters")
        
        # Verify structure
        assert "Parameter" in params_df.columns
        assert "Value" in params_df.columns
        assert len(params_df) == 4
        
        # Convert to dict for easier checking
        params_dict = dict(zip(params_df["Parameter"], params_df["Value"]))
        
        assert params_dict["Group By"] == "Company"
        assert params_dict["Value Column"] == "Revenue"
        assert params_dict["Time Column"] == "Date"
        assert params_dict["Thresholds"] == "[10, 20, 50]"
    
    def test_export_concentration_excel_without_details(self, temp_dir: Path):
        """Test Excel export when details are missing."""
        excel_path = temp_dir / "no_details.xlsx"
        
        results = {
            "group_by": "Company",
            "value_column": "Revenue",
            "by_period": [
                {
                    "period": "2023-Q1",
                    "total": 1000,
                    "top_10": {"count": 1, "value": 100, "pct_of_total": 10.0}
                }
            ]
        }
        
        ExportService.export_concentration_excel(results, excel_path)
        
        workbook = load_workbook(excel_path)
        sheet_names = workbook.sheetnames
        
        # Should have Summary and Parameters, but not Details
        assert "Summary" in sheet_names
        assert "Parameters" in sheet_names
        assert "Details" not in sheet_names
    
    def test_export_concentration_excel_with_formulas(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test Excel export with formulas enabled."""
        excel_path = temp_dir / "with_formulas.xlsx"
        
        result_path = ExportService.export_concentration_excel(
            sample_concentration_results,
            excel_path,
            include_formulas=True
        )
        
        assert result_path == str(excel_path)
        assert excel_path.exists()
        
        # For now, just verify the file is created correctly
        # Formula implementation is TODO in the code
        workbook = load_workbook(excel_path)
        assert len(workbook.sheetnames) >= 3
    
    def test_export_concentration_json(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test JSON export preserves all data."""
        json_path = temp_dir / "concentration_results.json"
        
        result_path = ExportService.export_concentration_json(
            sample_concentration_results,
            json_path
        )
        
        assert result_path == str(json_path)
        assert json_path.exists()
        
        # Read and verify JSON
        with open(json_path, "r") as f:
            exported_data = json.load(f)
        
        # Verify all major sections are preserved
        assert exported_data["group_by"] == sample_concentration_results["group_by"]
        assert exported_data["value_column"] == sample_concentration_results["value_column"]
        assert exported_data["time_column"] == sample_concentration_results["time_column"]
        assert exported_data["thresholds"] == sample_concentration_results["thresholds"]
        assert len(exported_data["by_period"]) == len(sample_concentration_results["by_period"])
        assert exported_data["details"] == sample_concentration_results["details"]
        
        # Verify specific period data
        q1_data = exported_data["by_period"][0]
        assert q1_data["period"] == "2023-Q1"
        assert q1_data["total"] == 2500000
        assert q1_data["top_10"]["count"] == 1
        assert q1_data["top_10"]["value"] == 1000000
        assert q1_data["top_10"]["pct_of_total"] == 40.0
    
    def test_csv_row_structure_validation(self, temp_dir: Path):
        """Test CSV export produces correct row structure."""
        csv_path = temp_dir / "structure_test.csv"
        
        # Create minimal test data
        results = {
            "by_period": [
                {
                    "period": "2023-Q1",
                    "top_10": {"count": 5, "value": 1000, "pct_of_total": 50.0},
                    "top_20": {"count": 10, "value": 1500, "pct_of_total": 75.0}
                },
                {
                    "period": "2023-Q2",
                    "top_10": {"count": 3, "value": 800, "pct_of_total": 40.0},
                    "top_50": {"count": 20, "value": 2000, "pct_of_total": 100.0}
                }
            ]
        }
        
        ExportService.export_concentration_csv(results, csv_path)
        
        df = pd.read_csv(csv_path)
        
        # Should have 4 rows total (Q1: 2 thresholds + Q2: 2 thresholds)
        assert len(df) == 4
        
        # Check Q1 data
        q1_rows = df[df["period"] == "2023-Q1"]
        assert len(q1_rows) == 2
        assert set(q1_rows["threshold"]) == {10, 20}
        
        # Check Q2 data
        q2_rows = df[df["period"] == "2023-Q2"]
        assert len(q2_rows) == 2
        assert set(q2_rows["threshold"]) == {10, 50}
    
    def test_export_with_empty_results(self, temp_dir: Path):
        """Test exports with empty or minimal results."""
        empty_results = {}
        
        # Test CSV
        csv_path = temp_dir / "empty.csv"
        ExportService.export_concentration_csv(empty_results, csv_path)
        assert csv_path.exists()
        
        # When there's no data, an empty DataFrame with no columns is created
        try:
            df = pd.read_csv(csv_path)
            assert len(df) == 0
        except pd.errors.EmptyDataError:
            # This is expected when the CSV has no columns
            pass
        
        # Test Excel
        excel_path = temp_dir / "empty.xlsx"
        ExportService.export_concentration_excel(empty_results, excel_path)
        workbook = load_workbook(excel_path)
        # Should at least have Parameters sheet
        assert "Parameters" in workbook.sheetnames
        
        # Test JSON
        json_path = temp_dir / "empty.json"
        ExportService.export_concentration_json(empty_results, json_path)
        with open(json_path, "r") as f:
            data = json.load(f)
        assert data == {}
    
    def test_missing_threshold_data(self, temp_dir: Path):
        """Test handling of missing threshold data in period."""
        csv_path = temp_dir / "missing_thresholds.csv"
        
        results = {
            "by_period": [
                {
                    "period": "2023-Q1",
                    "top_10": {"count": 1, "value": 100, "pct_of_total": 10.0}
                    # Missing top_20 and top_50
                }
            ]
        }
        
        ExportService.export_concentration_csv(results, csv_path)
        
        df = pd.read_csv(csv_path)
        assert len(df) == 1  # Only one threshold present
        assert df.iloc[0]["threshold"] == 10
        assert df.iloc[0]["count"] == 1
    
    def test_path_string_and_pathlib_compatibility(self, temp_dir: Path, sample_concentration_results: Dict[str, Any]):
        """Test that export functions accept both string and Path objects."""
        # Test with Path objects
        csv_path_obj = temp_dir / "test_path.csv"
        excel_path_obj = temp_dir / "test_path.xlsx"
        json_path_obj = temp_dir / "test_path.json"
        
        # Test with string paths
        csv_path_str = str(temp_dir / "test_string.csv")
        excel_path_str = str(temp_dir / "test_string.xlsx")
        json_path_str = str(temp_dir / "test_string.json")
        
        # All should work without error
        ExportService.export_concentration_csv(sample_concentration_results, csv_path_obj)
        ExportService.export_concentration_csv(sample_concentration_results, csv_path_str)
        
        ExportService.export_concentration_excel(sample_concentration_results, excel_path_obj)
        ExportService.export_concentration_excel(sample_concentration_results, excel_path_str)
        
        ExportService.export_concentration_json(sample_concentration_results, json_path_obj)
        ExportService.export_concentration_json(sample_concentration_results, json_path_str)
        
        # Verify all files were created
        assert Path(csv_path_obj).exists()
        assert Path(csv_path_str).exists()
        assert Path(excel_path_obj).exists()
        assert Path(excel_path_str).exists()
        assert Path(json_path_obj).exists()
        assert Path(json_path_str).exists()